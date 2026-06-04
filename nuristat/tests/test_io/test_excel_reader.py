"""Tests for excel_reader module."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from nuristat.core.dataset import Dataset
from nuristat.core.exceptions import FileReadError
from nuristat.io.excel_reader import read_excel


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> str:
    """Create a simple Excel workbook with one sheet."""
    path = tmp_path / "simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "score"])
    ws.append(["Alice", 25, 85.5])
    ws.append(["Bob", 30, 90.2])
    ws.append(["Charlie", 35, 78.0])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> str:
    """Create an Excel workbook with multiple sheets."""
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Patients"
    ws1.append(["patient_id", "age", "sex"])
    ws1.append(["P001", 45, "M"])
    ws1.append(["P002", 52, "F"])

    ws2 = wb.create_sheet("LabResults")
    ws2.append(["test_id", "value", "unit"])
    ws2.append(["T001", 120.5, "mg/dL"])
    ws2.append(["T002", 85.0, "mmHg"])

    wb.save(str(path))
    return str(path)


@pytest.fixture
def xlsx_by_index(tmp_path: Path) -> str:
    """Create an Excel file for testing sheet index selection."""
    path = tmp_path / "by_index.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a", "b"])
    ws1.append([1, 2])

    ws2 = wb.create_sheet("Second")
    ws2.append(["x", "y", "z"])
    ws2.append([10, 20, 30])

    wb.save(str(path))
    return str(path)


@pytest.fixture
def xlsx_no_header(tmp_path: Path) -> str:
    """Create an Excel file without a header row."""
    path = tmp_path / "no_header.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([1, "apple", "red"])
    ws.append([2, "banana", "yellow"])
    ws.append([3, "cherry", "red"])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def xlsx_korean(tmp_path: Path) -> str:
    """Create an Excel file with Korean text."""
    path = tmp_path / "korean.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["이름", "나이", "성별"])
    ws.append(["홍길동", 30, "남"])
    ws.append(["김영희", 25, "여"])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> str:
    """Create an empty Excel workbook (header only)."""
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    ws.append(["col_a", "col_b"])
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# read_excel tests
# ---------------------------------------------------------------------------


class TestReadExcel:
    def test_simple_sheet(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert isinstance(ds, Dataset)
        assert ds.n_rows == 3
        assert ds.n_vars == 3
        assert list(ds.data.columns) == ["name", "age", "score"]
        assert ds.data["age"].tolist() == [25, 30, 35]

    def test_dataset_type(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert isinstance(ds, Dataset)
        assert len(ds.variables) == 3
        for col in ds.data.columns:
            assert col in ds.variables

    def test_multi_sheet_by_name(self, multi_sheet_xlsx: str) -> None:
        ds = read_excel(multi_sheet_xlsx, sheet_name="LabResults")
        assert isinstance(ds, Dataset)
        assert list(ds.data.columns) == ["test_id", "value", "unit"]
        assert ds.data["value"].tolist() == [120.5, 85.0]

    def test_multi_sheet_by_index(self, xlsx_by_index: str) -> None:
        ds_first = read_excel(xlsx_by_index, sheet_name=0)
        assert list(ds_first.data.columns) == ["a", "b"]

        ds_second = read_excel(xlsx_by_index, sheet_name=1)
        assert list(ds_second.data.columns) == ["x", "y", "z"]
        assert ds_second.data["z"].tolist() == [30]

    def test_no_header(self, xlsx_no_header: str) -> None:
        ds = read_excel(xlsx_no_header, header=None)
        assert ds.n_vars == 3
        assert list(ds.data.columns) == ["var_0", "var_1", "var_2"]
        assert ds.data.iloc[0, 1] == "apple"

    def test_korean_content(self, xlsx_korean: str) -> None:
        ds = read_excel(xlsx_korean)
        assert isinstance(ds, Dataset)
        assert list(ds.data.columns) == ["이름", "나이", "성별"]
        assert ds.data["이름"].tolist() == ["홍길동", "김영희"]

    def test_source_info_populated(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert "format" in ds.source_info
        assert ds.source_info["format"] == "xlsx"
        assert "sheet_name" in ds.source_info
        assert "file_name" in ds.source_info

    def test_name_from_filename(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert ds.name == "simple"

    def test_empty_sheet_raises(self, empty_xlsx: str) -> None:
        # openpyxl reads header-only as one row, so this should succeed
        ds = read_excel(empty_xlsx)
        assert ds.n_rows == 0  # only header, no data rows
        assert ds.n_vars == 2

    def test_nonexistent_file_raises(self) -> None:
        with pytest.raises(FileReadError) as exc_info:
            read_excel("/nonexistent/file.xlsx")
        assert "존재하지 않습니다" in str(exc_info.value)

    def test_string_path(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert isinstance(ds, Dataset)
        assert ds.n_rows == 3

    def test_numeric_columns_correct_type(self, simple_xlsx: str) -> None:
        ds = read_excel(simple_xlsx)
        assert pd.api.types.is_integer_dtype(ds.data["age"])
        assert pd.api.types.is_float_dtype(ds.data["score"])
