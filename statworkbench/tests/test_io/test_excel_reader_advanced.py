"""Advanced tests for excel_reader module — targeting 80%+ coverage."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from statworkbench.core.dataset import Dataset
from statworkbench.core.exceptions import FileReadError
from statworkbench.io.excel_reader import read_excel


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def int_float_xlsx(tmp_path: Path) -> str:
    """Excel file with integer and float columns."""
    path = tmp_path / "int_float.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "value", "ratio"])
    for i in range(1, 6):
        ws.append([i, i * 10, i * 0.25])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def str_dtype_xlsx(tmp_path: Path) -> str:
    """Excel file with string/text columns."""
    path = tmp_path / "str_dtype.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Text"
    ws.append(["code", "label", "note"])
    ws.append(["A001", "Alpha", "first"])
    ws.append(["B002", "Beta", "second"])
    ws.append(["C003", "Gamma", "third"])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def datetime_xlsx(tmp_path: Path) -> str:
    """Excel file with datetime column (openpyxl native datetime)."""
    path = tmp_path / "datetime.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.append(["event", "recorded_at"])
    ws.append(["login", datetime(2024, 1, 15, 9, 0, 0)])
    ws.append(["logout", datetime(2024, 1, 15, 17, 30, 0)])
    ws.append(["error", datetime(2024, 2, 1, 12, 0, 0)])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def missing_values_xlsx(tmp_path: Path) -> str:
    """Excel file with various missing values."""
    path = tmp_path / "missing.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey"
    ws.append(["id", "name", "score", "group"])
    ws.append([1, "Alice", 85.5, "A"])
    ws.append([2, None, 90.0, "B"])
    ws.append([3, "Charlie", None, "A"])
    ws.append([4, "Diana", 78.0, None])
    ws.append([5, None, None, None])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def three_sheet_xlsx(tmp_path: Path) -> str:
    """Excel workbook with three named sheets."""
    path = tmp_path / "three_sheets.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Alpha"
    ws1.append(["x"])
    ws1.append([1])

    ws2 = wb.create_sheet("Beta")
    ws2.append(["y", "z"])
    ws2.append([10, 20])

    ws3 = wb.create_sheet("Gamma")
    ws3.append(["p", "q", "r"])
    ws3.append([100, 200, 300])

    wb.save(str(path))
    return str(path)


@pytest.fixture
def no_header_xlsx(tmp_path: Path) -> str:
    """Excel file with no column header (header=None)."""
    path = tmp_path / "noheader.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw"
    ws.append([10, 20, 30])
    ws.append([40, 50, 60])
    ws.append([70, 80, 90])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def header_row2_xlsx(tmp_path: Path) -> str:
    """Excel file where header is at row index 1 (second row)."""
    path = tmp_path / "header_row2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shifted"
    ws.append(["This is a metadata comment row"])
    ws.append(["col_a", "col_b", "col_c"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def large_xlsx(tmp_path: Path) -> str:
    """Excel file with 500 rows."""
    path = tmp_path / "large.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Big"
    ws.append(["index", "value_a", "value_b"])
    for i in range(500):
        ws.append([i, i * 2.5, i % 7])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def single_column_xlsx(tmp_path: Path) -> str:
    """Excel file with a single column."""
    path = tmp_path / "single_col.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OneCol"
    ws.append(["value"])
    for v in [10, 20, 30, 40, 50]:
        ws.append([v])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def mixed_types_xlsx(tmp_path: Path) -> str:
    """Excel file mixing int, float, str in one sheet."""
    path = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.append(["int_col", "float_col", "str_col", "bool_col"])
    ws.append([1, 1.1, "hello", True])
    ws.append([2, 2.2, "world", False])
    ws.append([3, 3.3, "test", True])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def empty_sheet_xlsx(tmp_path: Path) -> str:
    """Excel file with a completely empty sheet (no data at all)."""
    path = tmp_path / "empty_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    # Write nothing — truly empty
    wb.save(str(path))
    return str(path)


@pytest.fixture
def korean_headers_xlsx(tmp_path: Path) -> str:
    """Excel file with Korean column headers and data."""
    path = tmp_path / "korean_headers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "한국어"
    ws.append(["지역명", "인구수", "면적"])
    ws.append(["서울", 9700000, 605.6])
    ws.append(["부산", 3400000, 770.1])
    ws.append(["대구", 2400000, 883.5])
    wb.save(str(path))
    return str(path)


@pytest.fixture
def special_chars_xlsx(tmp_path: Path) -> str:
    """Excel file with special characters in data."""
    path = tmp_path / "special.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Special"
    ws.append(["col_a", "col_b"])
    ws.append(["hello world", "foo,bar"])
    ws.append(["tab\there", 'quote"here'])
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# dtype tests
# ---------------------------------------------------------------------------


class TestExcelDtypes:
    def test_integer_dtype(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert pd.api.types.is_integer_dtype(ds.data["id"])

    def test_float_dtype(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert pd.api.types.is_float_dtype(ds.data["ratio"])

    def test_string_dtype(self, str_dtype_xlsx: str) -> None:
        ds = read_excel(str_dtype_xlsx)
        assert ds.data["code"].dtype == object
        assert ds.data["label"].tolist() == ["Alpha", "Beta", "Gamma"]

    def test_datetime_dtype(self, datetime_xlsx: str) -> None:
        ds = read_excel(datetime_xlsx)
        # pandas read_excel converts datetime cells to datetime64
        assert pd.api.types.is_datetime64_any_dtype(ds.data["recorded_at"])

    def test_mixed_types_columns(self, mixed_types_xlsx: str) -> None:
        ds = read_excel(mixed_types_xlsx)
        assert pd.api.types.is_integer_dtype(ds.data["int_col"])
        assert pd.api.types.is_float_dtype(ds.data["float_col"])
        assert ds.data["str_col"].dtype == object


# ---------------------------------------------------------------------------
# Sheet selection tests
# ---------------------------------------------------------------------------


class TestExcelSheetSelection:
    def test_select_sheet_by_name_alpha(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name="Alpha")
        assert list(ds.data.columns) == ["x"]
        assert ds.data["x"].tolist() == [1]

    def test_select_sheet_by_name_beta(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name="Beta")
        assert list(ds.data.columns) == ["y", "z"]

    def test_select_sheet_by_name_gamma(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name="Gamma")
        assert ds.n_vars == 3
        assert ds.data["r"].tolist() == [300]

    def test_select_sheet_by_index_0(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name=0)
        assert list(ds.data.columns) == ["x"]

    def test_select_sheet_by_index_1(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name=1)
        assert list(ds.data.columns) == ["y", "z"]

    def test_select_sheet_by_index_2(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name=2)
        assert ds.n_vars == 3

    def test_invalid_sheet_name_raises(self, three_sheet_xlsx: str) -> None:
        with pytest.raises(FileReadError):
            read_excel(three_sheet_xlsx, sheet_name="NonExistent")

    def test_invalid_sheet_index_raises(self, three_sheet_xlsx: str) -> None:
        with pytest.raises(FileReadError):
            read_excel(three_sheet_xlsx, sheet_name=99)


# ---------------------------------------------------------------------------
# Header tests
# ---------------------------------------------------------------------------


class TestExcelHeader:
    def test_no_header_columns_renamed(self, no_header_xlsx: str) -> None:
        ds = read_excel(no_header_xlsx, header=None)
        assert ds.n_vars == 3
        # Dataset sanitizes integer columns to var_0, var_1, var_2
        assert list(ds.data.columns) == ["var_0", "var_1", "var_2"]

    def test_no_header_data_intact(self, no_header_xlsx: str) -> None:
        ds = read_excel(no_header_xlsx, header=None)
        assert ds.n_rows == 3

    def test_header_at_row_1(self, header_row2_xlsx: str) -> None:
        ds = read_excel(header_row2_xlsx, header=1)
        assert list(ds.data.columns) == ["col_a", "col_b", "col_c"]
        assert ds.n_rows == 2


# ---------------------------------------------------------------------------
# Missing values tests
# ---------------------------------------------------------------------------


class TestExcelMissingValues:
    def test_nan_in_name_column(self, missing_values_xlsx: str) -> None:
        ds = read_excel(missing_values_xlsx)
        assert ds.data["name"].isna().sum() == 2

    def test_nan_in_score_column(self, missing_values_xlsx: str) -> None:
        ds = read_excel(missing_values_xlsx)
        assert ds.data["score"].isna().sum() == 2

    def test_nan_in_group_column(self, missing_values_xlsx: str) -> None:
        ds = read_excel(missing_values_xlsx)
        assert ds.data["group"].isna().sum() == 2

    def test_all_missing_row_exists(self, missing_values_xlsx: str) -> None:
        ds = read_excel(missing_values_xlsx)
        last_row = ds.data.iloc[-1]
        assert pd.isna(last_row["name"])
        assert pd.isna(last_row["score"])
        assert pd.isna(last_row["group"])


# ---------------------------------------------------------------------------
# source_info tests
# ---------------------------------------------------------------------------


class TestExcelSourceInfo:
    def test_source_info_format(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.source_info["format"] == "xlsx"

    def test_source_info_file_name(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.source_info["file_name"] == "int_float.xlsx"

    def test_source_info_file_path_absolute(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert Path(ds.source_info["file_path"]).is_absolute()

    def test_source_info_file_size_positive(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.source_info["file_size"] > 0

    def test_source_info_sheet_name_default(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.source_info["sheet_name"] == 0

    def test_source_info_sheet_name_by_name(self, three_sheet_xlsx: str) -> None:
        ds = read_excel(three_sheet_xlsx, sheet_name="Beta")
        assert ds.source_info["sheet_name"] == "Beta"

    def test_source_info_header(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.source_info["header"] == 0

    def test_source_info_header_none(self, no_header_xlsx: str) -> None:
        ds = read_excel(no_header_xlsx, header=None)
        assert ds.source_info["header"] is None


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------


class TestExcelErrors:
    def test_file_not_found(self) -> None:
        with pytest.raises(FileReadError) as exc_info:
            read_excel("/no/such/file.xlsx")
        assert "존재하지 않습니다" in str(exc_info.value)

    def test_error_contains_filepath(self) -> None:
        target = "/no/such/file.xlsx"
        with pytest.raises(FileReadError) as exc_info:
            read_excel(target)
        assert "file.xlsx" in str(exc_info.value)

    def test_directory_path_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileReadError):
            read_excel(str(tmp_path))

    def test_empty_sheet_no_data(self, empty_sheet_xlsx: str) -> None:
        # A truly empty sheet returns an empty DataFrame — not an error
        ds = read_excel(empty_sheet_xlsx)
        assert ds.n_rows == 0

    def test_invalid_sheet_name_error_message(self, three_sheet_xlsx: str) -> None:
        with pytest.raises(FileReadError):
            read_excel(three_sheet_xlsx, sheet_name="DoesNotExist")


# ---------------------------------------------------------------------------
# Misc / edge case tests
# ---------------------------------------------------------------------------


class TestExcelMisc:
    def test_large_file_row_count(self, large_xlsx: str) -> None:
        ds = read_excel(large_xlsx)
        assert ds.n_rows == 500

    def test_large_file_col_count(self, large_xlsx: str) -> None:
        ds = read_excel(large_xlsx)
        assert ds.n_vars == 3

    def test_single_column_file(self, single_column_xlsx: str) -> None:
        ds = read_excel(single_column_xlsx)
        assert ds.n_vars == 1
        assert ds.n_rows == 5

    def test_dataset_name_is_stem(self, int_float_xlsx: str) -> None:
        ds = read_excel(int_float_xlsx)
        assert ds.name == "int_float"

    def test_korean_headers_preserved(self, korean_headers_xlsx: str) -> None:
        ds = read_excel(korean_headers_xlsx)
        assert "지역명" in ds.data.columns
        assert "인구수" in ds.data.columns
        assert ds.n_rows == 3

    def test_special_chars_in_data(self, special_chars_xlsx: str) -> None:
        ds = read_excel(special_chars_xlsx)
        assert ds.n_rows == 2
        assert "hello world" in ds.data["col_a"].tolist()

    def test_returns_dataset_instance(self, int_float_xlsx: str) -> None:
        result = read_excel(int_float_xlsx)
        assert isinstance(result, Dataset)

    def test_variables_metadata_created(self, str_dtype_xlsx: str) -> None:
        ds = read_excel(str_dtype_xlsx)
        for col in ds.data.columns:
            assert col in ds.variables

    def test_extra_kwargs_forwarded(self, int_float_xlsx: str) -> None:
        # usecols kwarg should filter columns
        ds = read_excel(int_float_xlsx, usecols=["id", "value"])
        assert ds.n_vars == 2
        assert "ratio" not in ds.data.columns

    def test_nrows_kwarg(self, large_xlsx: str) -> None:
        ds = read_excel(large_xlsx, nrows=10)
        assert ds.n_rows == 10


# ---------------------------------------------------------------------------
# Branch coverage — ValueError (invalid sheet triggers ValueError in pandas)
# ---------------------------------------------------------------------------


class TestExcelBranchCoverage:
    def test_sheet_list_returns_first_sheet(self, three_sheet_xlsx: str) -> None:
        """sheet_name as list makes pandas return a dict; code picks first sheet."""
        ds = read_excel(three_sheet_xlsx, sheet_name=["Alpha", "Beta"])
        # Should succeed and return a Dataset (first sheet)
        assert isinstance(ds, Dataset)

    def test_sheet_list_single_element(self, three_sheet_xlsx: str) -> None:
        """Single-element list still produces a dict; code should handle it."""
        ds = read_excel(three_sheet_xlsx, sheet_name=["Gamma"])
        assert isinstance(ds, Dataset)
        assert ds.n_vars == 3
